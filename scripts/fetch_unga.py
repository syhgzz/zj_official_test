# -*- coding: utf-8 -*-
"""
走航甲烷检测(UNGA)数据导出
输出: 燃气数据.xlsx
Sheet: 任务列表 / 轨迹采样点 / 泄露点
"""
import sys
import os
import time
from datetime import datetime
from calendar import monthrange
from openpyxl import Workbook

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from lib.api_client import APIClient
from config.config import config

OUTPUT = os.path.join(os.path.dirname(__file__), '..', '燃气数据.xlsx')
PAGE_SIZE = 2000
DELAY = 0.15


def monthly_ranges(start_year, end_year):
    for year in range(start_year, end_year + 1):
        for month in range(1, 13):
            s = datetime(year, month, 1, 0, 0, 0)
            e = datetime(year, month, monthrange(year, month)[1], 23, 59, 59)
            yield int(s.timestamp()) * 1000, int(e.timestamp()) * 1000


def ts_to_str(ms):
    try:
        return datetime.fromtimestamp(ms / 1000).strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        return ''


def paginate(client, path, params, list_key):
    items = []
    page = 1
    while True:
        p = {**params, 'pageNum': page, 'pageSize': PAGE_SIZE}
        res = client.request('GET', path, params=p)
        if not res or res.get('code') != 200:
            break
        data = res.get('data', {})
        batch = data.get(list_key, [])
        if isinstance(batch, list):
            items.extend(batch)
        total = data.get('total', 0)
        if page * PAGE_SIZE >= (total or 1):
            break
        page += 1
        time.sleep(DELAY)
    return items


def write_sheet(ws, rows, fieldnames):
    ws.append(list(fieldnames))
    for r in rows:
        ws.append([r.get(k, '') for k in fieldnames])


def main():
    client = APIClient(config.host, config.app_key, config.app_secret, config.timeout)

    # 数据容器
    task_rows = []       # Sheet 1: 任务列表
    sampling_rows = []   # Sheet 2: 轨迹采样点
    leak_rows = []       # Sheet 3: 泄露点
    processed_tasks = set()
    seen_leaks = set()

    print("=" * 60)
    print("燃气数据导出 (UNGA)")
    print("=" * 60)

    # 1. 按月获取泄露点
    print("获取泄露点...")
    for sm, em in monthly_ranges(2018, 2026):
        label = datetime.fromtimestamp(sm / 1000).strftime('%Y-%m')
        leaks = paginate(client, '/api/v1/unga/leaks',
                         {'startTime': sm, 'endTime': em}, 'leaks')
        new_count = 0
        for lk in leaks:
            lid = lk.get('leakId', '')
            if lid in seen_leaks:
                continue
            seen_leaks.add(lid)
            new_count += 1
            leak_rows.append({
                'leakId': lk.get('leakId', ''),
                'taskId': lk.get('taskId', ''),
                '经度': lk.get('longitude'),
                '纬度': lk.get('latitude'),
                '甲烷浓度': lk.get('gasConcentration'),
                '泄露等级': lk.get('level'),
                '泄露状态': lk.get('status'),
                '检测时间': ts_to_str(lk.get('detectionTime')),
                '附近管道': lk.get('nearbyPipeline', ''),
                'regionCode': lk.get('regionCode', ''),
            })
        if new_count:
            print(f"  [{label}] {len(leaks)} 条, 新增 {new_count} 个")

    print(f"  累计 {len(leak_rows)} 个泄露点\n")

    # 2. 按月获取任务, 对每个任务获取轨迹采样点
    for sm, em in monthly_ranges(2018, 2026):
        label = datetime.fromtimestamp(sm / 1000).strftime('%Y-%m')
        print(f"[{label}] 获取任务... ", end='', flush=True)
        tasks = paginate(client, '/api/v1/unga/tasks',
                         {'startTime': sm, 'endTime': em}, 'tasks')
        new_count = 0
        for t in tasks:
            tid = t.get('taskId', '')
            if not tid or tid in processed_tasks:
                continue
            processed_tasks.add(tid)
            new_count += 1

            # 任务基础信息
            task_rows.append({
                'taskId': tid,
                'taskName': t.get('taskName', ''),
                '任务开始时间': ts_to_str(t.get('startTime')),
                '任务结束时间': ts_to_str(t.get('endTime')),
                '里程(km)': t.get('mileage'),
                '疑似泄露数': t.get('suspectedLeakCount'),
                '平均甲烷浓度': t.get('avgGasConcentration'),
                '最大甲烷浓度': t.get('maxGasConcentration'),
                '任务状态': t.get('status'),
            })

            # 获取轨迹采样点
            path = f'/api/v1/unga/tasks/{tid}/trajectory'
            res = client.request('GET', path, params={})
            time.sleep(DELAY)
            if not res or res.get('code') != 200:
                continue

            data = res.get('data', {})
            if not isinstance(data, dict):
                continue
            traj = data.get('trajectory', {})
            if not isinstance(traj, dict):
                continue
            sps = traj.get('samplingPoints', [])
            if not isinstance(sps, list):
                continue

            for sp in sps:
                if not isinstance(sp, dict):
                    continue
                sampling_rows.append({
                    'taskId': tid,
                    '采样点序号': sp.get('sequence'),
                    '经度': sp.get('longitude'),
                    '纬度': sp.get('latitude'),
                    '甲烷浓度': sp.get('gasConcentration'),
                    '检测时间': ts_to_str(sp.get('timestamp')),
                })

        print(f"{len(tasks)} 个任务, 新 {new_count} 个")

    print(f"\n共 {len(task_rows)} 个任务, {len(sampling_rows)} 个采样点")

    # 写入 XLSX
    wb = Workbook()

    ws_tasks = wb.active
    ws_tasks.title = '任务列表'
    task_fields = ['taskId', 'taskName', '任务开始时间', '任务结束时间',
                   '里程(km)', '疑似泄露数', '平均甲烷浓度', '最大甲烷浓度', '任务状态']
    write_sheet(ws_tasks, task_rows, task_fields)

    ws_samples = wb.create_sheet('轨迹采样点')
    sample_fields = ['taskId', '采样点序号', '经度', '纬度', '甲烷浓度', '检测时间']
    write_sheet(ws_samples, sampling_rows, sample_fields)

    ws_leaks = wb.create_sheet('泄露点')
    leak_fields = ['leakId', 'taskId', '经度', '纬度', '甲烷浓度',
                   '泄露等级', '泄露状态', '检测时间', '附近管道', 'regionCode']
    write_sheet(ws_leaks, leak_rows, leak_fields)

    wb.save(OUTPUT)
    print(f"完成! Sheet1 {len(task_rows)} 行, Sheet2 {len(sampling_rows)} 行, Sheet3 {len(leak_rows)} 行")
    print(f"保存至: {OUTPUT}")


if __name__ == '__main__':
    main()
