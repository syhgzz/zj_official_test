# -*- coding: utf-8 -*-
"""
沉降态势感知(UPSS)数据导出
输出: 沉降数据.xlsx
Sheet: 沉降期次 / 沉降监测数据(regional+gridRate+gridGradient合并) / 预警记录
"""
import sys
import os
import time
from datetime import datetime
from openpyxl import Workbook

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from lib.api_client import APIClient
from config.config import config

OUTPUT = os.path.join(os.path.dirname(__file__), '..', '沉降数据.xlsx')
PAGE_SIZE = 1000
DELAY = 0.15


def ts_to_str(ms):
    try:
        return datetime.fromtimestamp(ms / 1000).strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        return ''


def write_sheet(ws, rows, fieldnames):
    ws.append(list(fieldnames))
    for r in rows:
        ws.append([r.get(k, '') for k in fieldnames])


def main():
    client = APIClient(config.host, config.app_key, config.app_secret, config.timeout)

    print("=" * 60)
    print("沉降数据导出 (UPSS)")
    print("=" * 60)

    # === 获取沉降期次 (不传时间范围以获取全部期次) ===
    print("获取沉降期次列表...")
    key_coords = {'minLng': 73, 'maxLng': 136, 'minLat': 3, 'maxLat': 54}
    res = client.request('GET', '/api/v1/upss/periods', params=key_coords)
    time.sleep(DELAY)
    all_periods = res.get('data', []) if res and res.get('code') == 200 else []
    issues = [p.get('issue') for p in all_periods
              if isinstance(p, dict) and p.get('issue')]
    print(f"  共 {len(issues)} 个沉降期次")

    # === 沉降监测数据 (regional + gridRate + gridGradient 按 pointCode+issue 合并) ===
    merged = {}  # key: (pointCode, issue) -> row

    for issue in issues:
        print(f"\n[{issue}] 获取数据...")

        # regional
        res_reg = client.request('GET', '/api/v1/upss/statistics/regional', params={
            'issue': issue, 'dimension': 'admin',
            'pageNum': 1, 'pageSize': PAGE_SIZE,
            **key_coords,
        })
        time.sleep(DELAY)
        if res_reg and res_reg.get('code') == 200:
            points = res_reg['data'].get('points', [])
            for pt in points:
                pc = pt.get('pointCode', '')
                key = (pc, issue)
                if key not in merged:
                    merged[key] = {'pointCode': pc, 'issue': issue}
                merged[key]['经度'] = pt.get('longitude')
                merged[key]['纬度'] = pt.get('latitude')
                merged[key]['累计沉降量'] = pt.get('subsidence')
                merged[key]['精度'] = pt.get('precision')
                ts = pt.get('timestamp')
                if ts:
                    merged[key]['数据时间'] = ts_to_str(ts)
            print(f"  regional: {len(points)} 点")

        # gridRate
        res_rate = client.request('GET', '/api/v1/upss/statistics/gridRate', params={
            'issue': issue,
            **key_coords,
        })
        time.sleep(DELAY)
        if res_rate and res_rate.get('code') == 200:
            points = res_rate['data'].get('points', [])
            for pt in points:
                pc = pt.get('pointCode', '')
                key = (pc, issue)
                if key not in merged:
                    merged[key] = {'pointCode': pc, 'issue': issue}
                    merged[key]['经度'] = pt.get('longitude')
                    merged[key]['纬度'] = pt.get('latitude')
                merged[key]['沉降速率'] = pt.get('value')
            print(f"  gridRate: {len(points)} 点")

        # gridGradient
        res_grad = client.request('GET', '/api/v1/upss/statistics/gridGradient', params={
            'issue': issue,
            **key_coords,
        })
        time.sleep(DELAY)
        if res_grad and res_grad.get('code') == 200:
            points = res_grad['data'].get('points', [])
            for pt in points:
                pc = pt.get('pointCode', '')
                key = (pc, issue)
                if key not in merged:
                    merged[key] = {'pointCode': pc, 'issue': issue}
                    merged[key]['经度'] = pt.get('longitude')
                    merged[key]['纬度'] = pt.get('latitude')
                merged[key]['沉降梯度'] = pt.get('value')
            print(f"  gridGradient: {len(points)} 点")

    # === 获取预警记录 ===
    print("\n获取预警信息...")
    res_warn = client.request('GET', '/api/v1/upss/visualization/warning/issue', params={
        'startTime': int(datetime(2018, 1, 1).timestamp() * 1000),
        'endTime': int(datetime(2026, 12, 31, 23, 59, 59).timestamp() * 1000),
        **key_coords,
    })
    time.sleep(DELAY)
    warnings = []
    if res_warn and res_warn.get('code') == 200:
        data = res_warn.get('data', {})
        issue = data.get('issue', '')
        issue_date = data.get('issueDate', '')
        for w in data.get('warnings', []):
            warnings.append({
                'issue': issue,
                'issueDate': issue_date,
                'pointCode': w.get('pointCode', ''),
                '经度': w.get('longitude'),
                '纬度': w.get('latitude'),
                '预警等级': w.get('level'),
                '累计沉降量': w.get('subsidence'),
                '沉降速率': w.get('rate'),
                '预警时间': ts_to_str(w.get('time')),
                'regionCode': w.get('regionCode', ''),
                'regionName': w.get('regionName', ''),
                '核查状态': w.get('checkState', ''),
            })
        print(f"  共 {len(warnings)} 条预警记录")

    # === 沉降态势统计(摘要) ===
    print("获取沉降态势统计...")
    summary_row = {}
    res_stat = client.request('GET', '/api/v1/upss/visualization/statistics/issue', params={
        'startTime': int(datetime(2018, 1, 1).timestamp() * 1000),
        'endTime': int(datetime(2026, 12, 31, 23, 59, 59).timestamp() * 1000),
        **key_coords,
    })
    time.sleep(DELAY)
    if res_stat and res_stat.get('code') == 200:
        sd = res_stat.get('data', {})
        if isinstance(sd, dict):
            summary_row = {
                'issue': sd.get('issue', ''),
                'totalPoints': sd.get('totalPoints', ''),
                'maxSubsidence': sd.get('maxSubsidence', ''),
                'avgSubsidence': sd.get('avgSubsidence', ''),
                'maxRate': sd.get('maxRate', ''),
                'regionCode': sd.get('regionCode', ''),
                'regionName': sd.get('regionName', ''),
            }

    # === 写入 XLSX ===
    wb = Workbook()

    # Sheet 1: 沉降期次
    ws_periods = wb.active
    ws_periods.title = '沉降期次'
    period_fields = ['issue', 'issueDate', 'prevIssue', 'daysInterval',
                     'totalPoints', 'warningPoints', 'maxSubsidence',
                     'avgSubsidence', 'regionCode', 'regionName']
    period_rows = []
    for p in all_periods:
        if isinstance(p, dict):
            row = {}
            for f in period_fields:
                row[f] = p.get(f, '')
            period_rows.append(row)
    if summary_row:
        smr = {'issue': summary_row.get('issue', ''), 'issueDate': '',
               'prevIssue': '', 'daysInterval': '',
               'totalPoints': summary_row.get('totalPoints', ''),
               'warningPoints': '', 'maxSubsidence': summary_row.get('maxSubsidence', ''),
               'avgSubsidence': summary_row.get('avgSubsidence', ''),
               'regionCode': summary_row.get('regionCode', ''),
               'regionName': summary_row.get('regionName', '')}
        period_rows.append(smr)
    write_sheet(ws_periods, period_rows, period_fields)

    # Sheet 2: 沉降监测数据
    ws_monitor = wb.create_sheet('沉降监测数据')
    monitor_fields = ['pointCode', 'issue', '经度', '纬度', '累计沉降量',
                      '沉降速率', '沉降梯度', '精度', '数据时间']
    write_sheet(ws_monitor, list(merged.values()), monitor_fields)

    # Sheet 3: 预警记录
    ws_warn = wb.create_sheet('预警记录')
    warn_fields = ['issue', 'issueDate', 'pointCode', '经度', '纬度',
                   '预警等级', '累计沉降量', '沉降速率', '预警时间',
                   'regionCode', 'regionName', '核查状态']
    write_sheet(ws_warn, warnings, warn_fields)

    wb.save(OUTPUT)
    print(f"\n完成! Sheet1 {len(period_rows)} 行, Sheet2 {len(merged)} 行, Sheet3 {len(warnings)} 行")
    print(f"保存至: {OUTPUT}")


if __name__ == '__main__':
    main()
